from __future__ import annotations

import html
import re
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("docs/HymnLink_data.xlsx")
SOP_CCLI_URL = "https://sop.org/copyright-ccli/"
HYMNARY_HOLC_URL = "https://hymnary.org/hymnal/HOLC1986"
TAIWANBIBLE_SEARCH_URL = "https://www.taiwanbible.com/web/search.jsp"


@dataclass
class SopRecord:
    title_zh: str
    title_en: str
    album: str
    series: str
    lyricist: str
    composer: str
    ccli: str
    copyright_text: str


@dataclass
class HymnaryRecord:
    number: str
    title_zh: str
    title_en: str
    tune: str
    href: str
    author: str = ""
    composer: str = ""
    copyright_text: str = ""


@dataclass
class TaiwanBibleRecord:
    title_zh: str
    album: str
    id: str


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_td = False
        self.current_cell: list[str] = []
        self.current_row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() == "td":
            self.in_td = True
            self.current_cell = []

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "td" and self.in_td:
            self.current_row.append(normalize_space("".join(self.current_cell)))
            self.in_td = False
        elif tag == "tr":
            if len(self.current_row) >= 8:
                self.rows.append(self.current_row[:8])
            self.current_row = []

    def handle_data(self, data: str) -> None:
        if self.in_td:
            self.current_cell.append(data)


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(value or "")).strip()


def title_key(value: str) -> str:
    value = normalize_space(value)
    value = re.sub(r"[\W_]+", "", value, flags=re.UNICODE)
    value = value.replace("祢", "你").replace("袮", "你")
    return value.lower()


def title_candidates(title: str) -> set[str]:
    title = normalize_space(title)
    candidates = {title_key(title)}
    for match in re.findall(r"[\[【](.*?)[\]】]", title):
        candidates.add(title_key(match))
    without_brackets = re.sub(r"[\[【].*?[\]】]", "", title)
    candidates.add(title_key(without_brackets))
    without_english_parentheses = re.sub(r"\([^)]*[A-Za-z][^)]*\)", "", title)
    candidates.add(title_key(without_english_parentheses))
    for part in re.split(r"[，,。；;！!？?：:\s]+", without_english_parentheses):
        if len(title_key(part)) >= 2:
            candidates.add(title_key(part))
    return {candidate for candidate in candidates if candidate}


def strip_tags(value: str) -> str:
    return normalize_space(re.sub(r"<.*?>", " ", value, flags=re.S))


def fetch_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        return response.read().decode("utf-8", "ignore")


def fetch_sop_records() -> list[SopRecord]:
    raw_html = fetch_text(SOP_CCLI_URL)

    parser = TableParser()
    parser.feed(raw_html)

    records: list[SopRecord] = []
    for row in parser.rows:
        records.append(
            SopRecord(
                title_zh=row[0],
                title_en=row[1],
                album=row[2],
                series=row[3],
                lyricist=row[4],
                composer=row[5],
                ccli=row[6],
                copyright_text=row[7],
            )
        )
    return records


def fetch_hymnary_records() -> list[HymnaryRecord]:
    records: list[HymnaryRecord] = []

    for page in range(6):
        url = HYMNARY_HOLC_URL if page == 0 else f"{HYMNARY_HOLC_URL}?page={page}"
        raw_html = fetch_text(url)
        rows = re.findall(r"<tr class='result-row[^']*'>(.*?)</tr>", raw_html, flags=re.S)
        for row in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, flags=re.S)
            if len(cells) < 3:
                continue
            number = strip_tags(cells[0])
            title = strip_tags(cells[1])
            tune = strip_tags(cells[2])
            hrefs = re.findall(r"href=['\"]([^'\"]+)['\"]", cells[1])
            english_match = re.search(r"\(([^()]*)\)\s*$", title)
            english_title = normalize_space(english_match.group(1) if english_match else "")
            chinese_title = normalize_space(re.sub(r"\([^)]*[A-Za-z][^)]*\)\s*$", "", title))

            records.append(
                HymnaryRecord(
                    number=number,
                    title_zh=chinese_title,
                    title_en=english_title,
                    tune=tune,
                    href=hrefs[0] if hrefs else f"/hymn/HOLC1986/{number}",
                )
            )

    return records


def enrich_hymnary_detail(record: HymnaryRecord) -> HymnaryRecord:
    url = f"https://hymnary.org{record.href}" if record.href.startswith("/") else record.href
    raw_html = fetch_text(url)

    def label(name: str) -> str:
        pattern = rf"<b>{re.escape(name)}:</b>\s*</td>\s*<td>(.*?)</td>"
        match = re.search(pattern, raw_html, flags=re.S)
        return strip_tags(match.group(1)) if match else ""

    text_value = label("Text")
    if not record.title_en:
        english_match = re.search(r"\(([^()]*)\)", text_value)
        if english_match:
            record.title_en = normalize_space(english_match.group(1))

    record.author = label("Author")
    record.composer = label("Composer")
    record.copyright_text = label("Copyright")
    return record


def fetch_taiwanbible_record(title: str) -> TaiwanBibleRecord | None:
    query = urllib.parse.urlencode({"area": "lyrics", "keyword": title})
    raw_html = fetch_text(f"{TAIWANBIBLE_SEARCH_URL}?{query}")
    rows = re.findall(r"<tr>\s*<td width=\"120\">.*?</tr>", raw_html, flags=re.S)

    for row in rows:
        link_match = re.search(r'<a href="/web/lyrics/view\.jsp\?ID=(\d+)">([^<]+)</a>', row)
        album_match = re.search(r"\(專輯:([^)]+)\)", row)
        if not link_match or not album_match:
            continue

        result_title = normalize_space(link_match.group(2))
        if title_key(result_title) != title_key(title):
            continue

        song_id = link_match.group(1)
        album = fetch_taiwanbible_album(song_id) or normalize_space(album_match.group(1))
        return TaiwanBibleRecord(title_zh=result_title, album=album, id=song_id)

    return None


def fetch_taiwanbible_album(song_id: str) -> str:
    raw_html = fetch_text(f"https://www.taiwanbible.com/web/lyrics/view.jsp?ID={song_id}")
    match = re.search(r"<p>專輯:\s*<a [^>]*>(.*?)</a></p>", raw_html, flags=re.S)
    return strip_tags(match.group(1)) if match else ""


def suspicious_album(value: str) -> bool:
    value = normalize_space(value)
    return bool(value) and value.count("(") != value.count(")")


def split_pipe(value: str) -> list[str]:
    return [normalize_space(part) for part in re.split(r"[|｜]", value or "") if normalize_space(part)]


def merge_tags(existing: str, inferred: str) -> str:
    tags: list[str] = []
    for tag in split_pipe(existing):
        if not suspicious_album(tag):
            tags.append(tag)
    for tag in split_pipe(inferred):
        tags.append(tag)
    return "|".join(dict.fromkeys(tags))


def infer_categories(title: str, album: str = "", series: str = "") -> str:
    text = f"{title} {album} {series}"
    categories: list[str] = []

    rules = [
        ("kids", ["兒童", "孩子", "小小", "小門徒"]),
        ("christmas", ["聖誕", "降生", "馬槽", "以馬內利"]),
        ("easter", ["復活", "得勝", "十架", "十字架"]),
        ("communion", ["聖餐", "寶血", "羔羊"]),
        ("thanksgiving", ["感恩", "稱謝", "謝謝", "數算主恩"]),
        ("prayer", ["禱告", "祈求", "懇求", "呼求"]),
        ("praise", ["讚美", "頌讚", "歡呼", "稱頌", "歌唱", "高舉", "榮耀歸"]),
        ("worship", ["敬拜", "獻上", "愛祢", "愛你", "榮耀", "聖潔"]),
        ("gospel", ["福音", "救恩", "救主", "十架", "耶穌愛你"]),
        ("devotional", ["跟隨", "親近", "尋求", "仰望", "渴慕", "一生", "生命"]),
        ("meditation", ["安靜", "默想", "平安", "看顧", "等候"]),
        ("hymn", ["聖詩", "真神", "有福的確據", "三一頌"]),
    ]

    for code, needles in rules:
        if any(needle in text for needle in needles):
            categories.append(code)

    if not categories:
        categories = ["worship"]

    return "|".join(dict.fromkeys(categories))


def infer_language(title: str, series: str = "") -> str:
    if "台語" in series or "閩南" in series or "疼" in title or "台語" in title:
        return "Taiwanese"
    if re.search(r"[A-Za-z]", title):
        return "English"
    return "Mandarin"


def infer_tags(title: str, album: str = "", series: str = "") -> str:
    terms = []
    source = f"{title} {album} {series}"
    tag_rules = {
        "耶穌": "耶穌",
        "敬拜": "敬拜",
        "讚美": "讚美",
        "禱告": "禱告",
        "感恩": "感恩",
        "聖潔": "聖潔",
        "榮耀": "榮耀",
        "十架": "十架",
        "十字架": "十字架",
        "復活": "復活",
        "聖誕": "聖誕",
        "平安": "平安",
        "愛": "愛",
        "信心": "信心",
        "恩典": "恩典",
        "兒童": "兒童",
        "台語": "台語",
        "聖靈": "聖靈",
        "醫治": "醫治",
        "宣教": "宣教",
        "家庭": "家庭",
        "祝福": "祝福",
        "喜樂": "喜樂",
        "自由": "自由",
        "創造": "創造",
        "跟隨": "跟隨",
        "應許": "應許",
    }
    for needle, tag in tag_rules.items():
        if needle in source:
            terms.append(tag)
    if album:
        terms.append(album)
    if series:
        terms.append(series)
    return "|".join(dict.fromkeys(terms))


def copyright_parts(copyright_text: str) -> tuple[str, str]:
    match = re.match(r"(\d{4})\s*(.*)", copyright_text)
    if not match:
        return "", copyright_text
    return match.group(1), match.group(2).strip()


def main() -> None:
    sop_records = fetch_sop_records()
    sop_lookup: dict[str, SopRecord] = {}
    for record in sop_records:
        for candidate in title_candidates(record.title_zh):
            sop_lookup.setdefault(candidate, record)

    hymnary_records = fetch_hymnary_records()
    hymnary_lookup: dict[str, HymnaryRecord] = {}
    for record in hymnary_records:
        for candidate in title_candidates(record.title_zh):
            hymnary_lookup.setdefault(candidate, record)

    workbook = load_workbook(WORKBOOK)
    sheet = workbook["Hymns"]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    col = {header: index + 1 for index, header in enumerate(headers)}

    matched = 0
    hymnary_matched = 0
    taiwanbible_matched = 0
    metadata_updates = 0
    heuristic_updates = 0

    for row in range(2, sheet.max_row + 1):
        title = normalize_space(str(sheet.cell(row, col["name_zh"]).value or ""))
        if not title or title.startswith("←"):
            continue

        sop_record = sop_lookup.get(title_key(title))
        hymnary_record = hymnary_lookup.get(title_key(title))
        taiwanbible_record = None

        if sop_record:
            matched += 1
            year, holder = copyright_parts(sop_record.copyright_text)
            updates = {
                "name_en": sop_record.title_en,
                "album": sop_record.album,
                "composer": sop_record.composer,
                "lyricist": sop_record.lyricist,
                "copyright_year": year,
                "copyright_holder": holder,
                "license": f"CCLI {sop_record.ccli}" if sop_record.ccli and sop_record.ccli != "Public Domain" else sop_record.ccli,
                "publisher": "Stream of Praise Music",
            }
            for field, value in updates.items():
                if value and not sheet.cell(row, col[field]).value:
                    sheet.cell(row, col[field]).value = value
                    metadata_updates += 1
        elif hymnary_record:
            hymnary_matched += 1
            detailed = enrich_hymnary_detail(hymnary_record)
            updates = {
                "name_en": detailed.title_en,
                "album": "生命聖詩 - Hymns of Life, 1986",
                "composer": detailed.composer,
                "lyricist": detailed.author,
                "publisher": "Chinese Alliance Press",
                "license": detailed.copyright_text or f"Hymnary HOLC1986 #{detailed.number}",
            }
            for field, value in updates.items():
                if value and not sheet.cell(row, col[field]).value:
                    sheet.cell(row, col[field]).value = value
                    metadata_updates += 1
            time.sleep(0.1)
        elif not sheet.cell(row, col["album"]).value or suspicious_album(str(sheet.cell(row, col["album"]).value or "")):
            taiwanbible_record = fetch_taiwanbible_record(title)
            if taiwanbible_record:
                taiwanbible_matched += 1
                sheet.cell(row, col["album"]).value = taiwanbible_record.album
                metadata_updates += 1
            time.sleep(0.05)

        album = normalize_space(str(sheet.cell(row, col["album"]).value or ""))
        series = sop_record.series if sop_record else ""

        if not sheet.cell(row, col["categories"]).value:
            sheet.cell(row, col["categories"]).value = infer_categories(title, album, series)
            heuristic_updates += 1
        if not sheet.cell(row, col["language"]).value:
            sheet.cell(row, col["language"]).value = infer_language(title, series)
            heuristic_updates += 1
        current_tags = str(sheet.cell(row, col["tags"]).value or "")
        new_tags = merge_tags(current_tags, infer_tags(title, album, series))
        if new_tags and new_tags != current_tags:
            sheet.cell(row, col["tags"]).value = new_tags
            heuristic_updates += 1

    workbook.save(WORKBOOK)
    print(f"SOP records parsed: {len(sop_records)}")
    print(f"Hymnary HOLC records parsed: {len(hymnary_records)}")
    print(f"Hymn rows matched to SOP metadata: {matched}")
    print(f"Hymn rows matched to Hymnary metadata: {hymnary_matched}")
    print(f"Hymn rows matched to TaiwanBible album metadata: {taiwanbible_matched}")
    print(f"Verified metadata cells filled: {metadata_updates}")
    print(f"Heuristic category/language/tag cells filled: {heuristic_updates}")
    print("Lyrics and YouTube URLs were not auto-filled; they require licensed lyrics and manual/official video verification.")


if __name__ == "__main__":
    main()
